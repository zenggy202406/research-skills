#!/usr/bin/env python3
"""Add a 'Selected' sheet to a paper-skimmer spreadsheet with scored/filtered papers.

Creates (or replaces) a 'Selected' sheet sorted by relevance score, color-coded by tier.
Preserves the original 'Papers' sheet untouched.
"""

import argparse
import json
import sys
import os

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print(json.dumps({"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl --break-system-packages"}))
    sys.exit(1)

HEADERS = ["Author(s)", "Year", "Title", "Research Question / Purpose",
           "Method Summary", "Key Findings", "Score", "Tier", "Rationale"]
COL_WIDTHS = [25, 8, 40, 35, 40, 45, 8, 12, 50]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
SCORE_ALIGN = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

TIER_FILLS = {
    "High": PatternFill("solid", fgColor="E2EFDA"),
    "Moderate": PatternFill("solid", fgColor="FFF2CC"),
    "Low": PatternFill("solid", fgColor="FCE4EC"),
}
TIER_FONTS = {
    "High": Font(name="Arial", size=10, bold=True, color="375623"),
    "Moderate": Font(name="Arial", size=10, bold=True, color="7F6000"),
    "Low": Font(name="Arial", size=10, color="C00000"),
}
SCORE_FONT_HIGH = Font(name="Arial", size=12, bold=True, color="375623")
SCORE_FONT_MOD = Font(name="Arial", size=12, bold=True, color="7F6000")
SCORE_FONT_LOW = Font(name="Arial", size=12, color="C00000")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Path to existing project .xlsx")
    parser.add_argument("--selections", required=True, help="JSON array of selection objects")
    parser.add_argument("--domain", default="", help="Research domain for header")
    parser.add_argument("--idea", default="", help="Core research idea for header")
    parser.add_argument("--rqs", default="", help="Research questions for header")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(json.dumps({"status": "error", "message": f"File not found: {args.file}"}))
        sys.exit(1)

    selections = json.loads(args.selections)
    selections.sort(key=lambda x: x.get("score", 0), reverse=True)

    wb = load_workbook(args.file)

    # Remove existing Selected sheet if present
    if "Selected" in wb.sheetnames:
        del wb["Selected"]

    ws = wb.create_sheet("Selected")
    ws.sheet_properties.tabColor = "1F4E79"

    current_row = 1

    # Research profile header block
    if args.domain or args.idea or args.rqs:
        profile_fill = PatternFill("solid", fgColor="D6E4F0")
        profile_font = Font(name="Arial", size=10, italic=True, color="1F4E79")
        profile_bold = Font(name="Arial", size=10, bold=True, color="1F4E79")

        if args.domain:
            ws.cell(row=current_row, column=1, value="Domain:").font = profile_bold
            ws.cell(row=current_row, column=1).fill = profile_fill
            c = ws.cell(row=current_row, column=2, value=args.domain)
            c.font = profile_font
            c.fill = profile_fill
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
            current_row += 1

        if args.idea:
            ws.cell(row=current_row, column=1, value="Core Idea:").font = profile_bold
            ws.cell(row=current_row, column=1).fill = profile_fill
            c = ws.cell(row=current_row, column=2, value=args.idea)
            c.font = profile_font
            c.fill = profile_fill
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
            current_row += 1

        if args.rqs:
            ws.cell(row=current_row, column=1, value="Research Qs:").font = profile_bold
            ws.cell(row=current_row, column=1).fill = profile_fill
            c = ws.cell(row=current_row, column=2, value=args.rqs)
            c.font = profile_font
            c.fill = profile_fill
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
            current_row += 1

        current_row += 1  # blank row

    # Column headers
    header_row = current_row
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[header_row].height = 30
    ws.freeze_panes = f"A{header_row + 1}"

    # Data rows
    high_count = mod_count = low_count = 0
    for i, sel in enumerate(selections):
        row = header_row + 1 + i
        tier = sel.get("tier", "Low")

        if tier == "High":
            high_count += 1
        elif tier == "Moderate":
            mod_count += 1
        else:
            low_count += 1

        row_fill = TIER_FILLS.get(tier, TIER_FILLS["Low"])

        values = [
            sel.get("authors", ""),
            sel.get("year", ""),
            sel.get("title", ""),
            sel.get("rq", ""),
            sel.get("method", ""),
            sel.get("findings", ""),
            sel.get("score", 0),
            tier,
            sel.get("rationale", ""),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.fill = row_fill

            if col_idx == 7:  # Score column
                cell.alignment = SCORE_ALIGN
                if tier == "High":
                    cell.font = SCORE_FONT_HIGH
                elif tier == "Moderate":
                    cell.font = SCORE_FONT_MOD
                else:
                    cell.font = SCORE_FONT_LOW
            elif col_idx == 8:  # Tier column
                cell.font = TIER_FONTS.get(tier, CELL_FONT)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGN

        ws.row_dimensions[row].height = 60

    # Autofilter on data range
    last_row = header_row + len(selections)
    ws.auto_filter.ref = f"A{header_row}:I{last_row}"

    wb.save(args.file)

    result = {
        "status": "success",
        "file": args.file,
        "total_scored": len(selections),
        "high_relevance": high_count,
        "moderate_relevance": mod_count,
        "low_relevance": low_count,
        "selected_count": high_count + mod_count,
    }
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
