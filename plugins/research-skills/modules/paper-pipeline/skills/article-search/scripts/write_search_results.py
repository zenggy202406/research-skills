#!/usr/bin/env python3
"""Write user-approved search results to a per-project Excel spreadsheet.

Creates the file with formatted headers if it doesn't exist;
appends new rows if it does. Warns on duplicate titles.
Accepts a JSON array of paper objects via --data.
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print(json.dumps({
        "status": "error",
        "message": "openpyxl not installed. Run: pip install openpyxl --break-system-packages"
    }))
    sys.exit(1)

HEADERS = [
    "Title",
    "Author(s)",
    "Year",
    "Citations",
    "Abstract",
    "URL",
    "Relevance",
    "Search Context",
]
COL_WIDTHS = [45, 25, 8, 10, 55, 30, 12, 35]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)

URL_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Relevance-based row highlights
RELEVANCE_FILLS = {
    "High": PatternFill("solid", fgColor="E2EFDA"),      # light green
    "Moderate": PatternFill("solid", fgColor="FFF2CC"),   # light yellow
}
DEFAULT_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")
DEFAULT_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")


def create_workbook(filepath: str) -> Workbook:
    """Create a new workbook with formatted headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Search Results"
    ws.sheet_properties.tabColor = "548235"

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(HEADERS))}1"
    wb.save(filepath)
    return wb


def check_duplicate(ws, title: str) -> bool:
    """Check if a paper with the same title already exists."""
    title_lower = title.strip().lower()
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0] and str(row[0]).strip().lower() == title_lower:
            return True
    return False


def append_row(ws, paper: dict, next_row: int):
    """Append a single paper row with formatting."""
    values = [
        paper.get("title", ""),
        paper.get("authors", ""),
        paper.get("year", ""),
        paper.get("citations", ""),
        paper.get("abstract", ""),
        paper.get("url", ""),
        paper.get("relevance", ""),
        paper.get("search_context", ""),
    ]

    relevance = str(paper.get("relevance", "")).strip()
    if relevance in RELEVANCE_FILLS:
        row_fill = RELEVANCE_FILLS[relevance]
    elif next_row % 2 == 0:
        row_fill = DEFAULT_FILL_EVEN
    else:
        row_fill = DEFAULT_FILL_ODD

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.font = CELL_FONT
        cell.alignment = CELL_ALIGN
        cell.border = THIN_BORDER
        cell.fill = row_fill

        # URL column gets hyperlink styling
        if col_idx == 6 and val:
            cell.font = URL_FONT

    ws.row_dimensions[next_row].height = 60


def main():
    parser = argparse.ArgumentParser(
        description="Write search results to project spreadsheet"
    )
    parser.add_argument(
        "--file", required=True, help="Path to the .xlsx file"
    )
    parser.add_argument(
        "--data", required=True,
        help='JSON array of paper objects, e.g. [{"title":"...","authors":"...","year":2023,"citations":45,"abstract":"...","url":"...","relevance":"High","search_context":"..."}]'
    )
    args = parser.parse_args()

    filepath = args.file

    # Parse JSON data
    try:
        papers = json.loads(args.data)
    except json.JSONDecodeError as e:
        print(json.dumps({
            "status": "error",
            "message": f"Invalid JSON in --data: {e}"
        }))
        sys.exit(1)

    if not isinstance(papers, list):
        print(json.dumps({
            "status": "error",
            "message": "--data must be a JSON array of paper objects"
        }))
        sys.exit(1)

    if not papers:
        print(json.dumps({
            "status": "error",
            "message": "No papers provided in --data"
        }))
        sys.exit(1)

    # Create or load workbook
    is_new = not os.path.exists(filepath)
    if is_new:
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        wb = create_workbook(filepath)
    else:
        wb = load_workbook(filepath)

    ws = wb["Search Results"] if "Search Results" in wb.sheetnames else wb.active

    # Process each paper
    added = []
    duplicates = []
    next_row = ws.max_row + 1

    for paper in papers:
        title = paper.get("title", "").strip()
        if not title:
            continue

        is_dup = False
        if not is_new:
            is_dup = check_duplicate(ws, title)

        if is_dup:
            duplicates.append(title)
            # Still add — user explicitly approved it — but warn

        append_row(ws, paper, next_row)
        added.append(title)
        next_row += 1

    # Update autofilter range
    last_col = chr(64 + len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{next_row - 1}"

    wb.save(filepath)

    total_papers = next_row - 1 - 1  # subtract header row

    result = {
        "status": "success",
        "file": filepath,
        "new_file": is_new,
        "papers_added": len(added),
        "total_papers": total_papers,
        "added_titles": added,
    }
    if duplicates:
        result["duplicate_warnings"] = duplicates
        result["warning"] = f"{len(duplicates)} paper(s) already existed in the spreadsheet. Added anyway since user approved."

    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
