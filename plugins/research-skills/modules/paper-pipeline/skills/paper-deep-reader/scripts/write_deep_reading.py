#!/usr/bin/env python3
"""Write a deep-reading analysis spreadsheet for a set of papers.

Creates a professionally formatted xlsx with:
  - Identification columns (blue): Authors, Year, Title, Journal
  - Integrated Summary column (green): A 300-word paragraph per paper integrating
    theory, methodology, findings, critical analysis, and user insights
"""

import argparse
import json
import os
import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl --break-system-packages"}))
    sys.exit(1)

# Column definitions: (field_key, header_label, width, group)
COLUMNS = [
    ("authors", "Author(s)",           22, "id"),
    ("year",    "Year",                 7, "id"),
    ("title",   "Title",              40, "id"),
    ("journal", "Journal / Source",    30, "id"),
    ("summary", "Integrated Summary",  90, "summary"),
]

# Group colors
GROUP_COLORS = {
    "id":      {"header": "2F5496", "cat": "D6E4F0", "font": "FFFFFF"},
    "summary": {"header": "375623", "cat": "E2EFDA", "font": "FFFFFF"},
}

GROUP_LABELS = {
    "id":      "Identification",
    "summary": "Integrated Summary",
}

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def create_workbook(filepath, papers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Deep Reading"
    ws.sheet_properties.tabColor = "375623"

    # Row 1: Category group headers (merged)
    group_spans = []
    current_group = None
    start_col = 1
    for i, (_, _, _, group) in enumerate(COLUMNS):
        col = i + 1
        if group != current_group:
            if current_group is not None:
                group_spans.append((current_group, start_col, col - 1))
            current_group = group
            start_col = col
    group_spans.append((current_group, start_col, len(COLUMNS)))

    for group, scol, ecol in group_spans:
        colors = GROUP_COLORS[group]
        label = GROUP_LABELS[group]
        cell = ws.cell(row=1, column=scol, value=label)
        cell.font = Font(name="Arial", bold=True, size=11, color=colors["font"])
        cell.fill = PatternFill("solid", fgColor=colors["cat"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if scol != ecol:
            ws.merge_cells(start_row=1, start_column=scol, end_row=1, end_column=ecol)
            for c in range(scol + 1, ecol + 1):
                bc = ws.cell(row=1, column=c)
                bc.fill = PatternFill("solid", fgColor=colors["cat"])
                bc.border = THIN_BORDER

    ws.row_dimensions[1].height = 25

    # Row 2: Column headers
    for col_idx, (_, label, width, group) in enumerate(COLUMNS, start=1):
        colors = GROUP_COLORS[group]
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(name="Arial", bold=True, size=10, color=colors["font"])
        cell.fill = PatternFill("solid", fgColor=colors["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 35
    ws.freeze_panes = "A3"

    # Data rows
    write_data_rows(ws, papers, start_row=3)

    # Autofilter
    last_row = 2 + len(papers)
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A2:{last_col}{last_row}"

    wb.save(filepath)
    return len(papers)


def write_data_rows(ws, papers, start_row):
    id_font = Font(name="Arial", size=9)
    summary_font = Font(name="Arial", size=9)
    id_align = Alignment(vertical="top", wrap_text=True)
    summary_align = Alignment(vertical="top", wrap_text=True)

    alt_fills = [
        PatternFill("solid", fgColor="FFFFFF"),
        PatternFill("solid", fgColor="F7F7F7"),
    ]

    for i, paper in enumerate(papers):
        row = start_row + i
        row_fill = alt_fills[i % 2]

        for col_idx, (key, _, _, group) in enumerate(COLUMNS, start=1):
            val = paper.get(key, "")
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.fill = row_fill

            if group == "summary":
                cell.font = summary_font
                cell.alignment = summary_align
            else:
                cell.font = id_font
                cell.alignment = id_align

        # Taller rows to accommodate the paragraph summary
        ws.row_dimensions[row].height = 200


def append_to_existing(filepath, papers):
    wb = load_workbook(filepath)
    ws = wb["Deep Reading"] if "Deep Reading" in wb.sheetnames else wb.active
    next_row = ws.max_row + 1
    write_data_rows(ws, papers, start_row=next_row)

    last_row = next_row + len(papers) - 1
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A2:{last_col}{last_row}"

    wb.save(filepath)

    # Count total data rows (subtract header rows)
    return ws.max_row - 2


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Output .xlsx path")
    parser.add_argument("--papers", required=True, help="JSON array of paper objects")
    args = parser.parse_args()

    papers = json.loads(args.papers)
    is_new = not os.path.exists(args.file)

    if is_new:
        os.makedirs(os.path.dirname(args.file) or ".", exist_ok=True)
        total = create_workbook(args.file, papers)
    else:
        total = append_to_existing(args.file, papers)

    result = {
        "status": "success",
        "file": args.file,
        "new_file": is_new,
        "papers_added": len(papers),
        "total_papers": total,
    }
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
