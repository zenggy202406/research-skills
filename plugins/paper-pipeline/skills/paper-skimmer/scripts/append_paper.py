#!/usr/bin/env python3
"""Append a paper's extracted info to a per-project Excel spreadsheet.

Creates the file with formatted headers if it doesn't exist;
appends a new row if it does. Warns on duplicate titles.
"""

import argparse
import os
import sys
import json
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print(json.dumps({"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl --break-system-packages"}))
    sys.exit(1)

HEADERS = ["Author(s)", "Year", "Title", "Research Question / Purpose", "Method Summary", "Key Findings"]
COL_WIDTHS = [25, 8, 45, 40, 45, 50]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def create_workbook(filepath: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Papers"
    ws.sheet_properties.tabColor = "2F5496"

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F1"
    wb.save(filepath)
    return wb


def check_duplicate(ws, title: str) -> bool:
    title_lower = title.strip().lower()
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        if row[0] and str(row[0]).strip().lower() == title_lower:
            return True
    return False


def append_row(ws, data: dict, next_row: int):
    values = [
        data["authors"],
        data["year"],
        data["title"],
        data["rq"],
        data["method"],
        data["findings"],
    ]
    even_row = next_row % 2 == 0
    row_fill = PatternFill("solid", fgColor="F2F2F2") if even_row else PatternFill("solid", fgColor="FFFFFF")

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.font = CELL_FONT
        cell.alignment = CELL_ALIGN
        cell.border = THIN_BORDER
        cell.fill = row_fill

    ws.row_dimensions[next_row].height = 60


def main():
    parser = argparse.ArgumentParser(description="Append paper info to project spreadsheet")
    parser.add_argument("--file", required=True, help="Path to the .xlsx file")
    parser.add_argument("--authors", required=True)
    parser.add_argument("--year", required=True, type=int)
    parser.add_argument("--title", required=True)
    parser.add_argument("--rq", required=True, help="Research question or purpose")
    parser.add_argument("--method", required=True, help="Method summary")
    parser.add_argument("--findings", required=True, help="Key findings")
    args = parser.parse_args()

    filepath = args.file
    is_new = not os.path.exists(filepath)

    if is_new:
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        wb = create_workbook(filepath)
    else:
        wb = load_workbook(filepath)

    ws = wb["Papers"] if "Papers" in wb.sheetnames else wb.active

    duplicate = False
    if not is_new:
        duplicate = check_duplicate(ws, args.title)

    next_row = ws.max_row + 1
    data = {
        "authors": args.authors,
        "year": args.year,
        "title": args.title,
        "rq": args.rq,
        "method": args.method,
        "findings": args.findings,
    }
    append_row(ws, data, next_row)

    # Update autofilter range
    ws.auto_filter.ref = f"A1:F{next_row}"

    wb.save(filepath)

    total_papers = next_row - 1
    result = {
        "status": "success",
        "file": filepath,
        "new_file": is_new,
        "total_papers": total_papers,
        "duplicate_warning": duplicate,
        "paper_added": args.title,
    }
    if duplicate:
        result["warning"] = f"A paper with title '{args.title}' already exists in the spreadsheet. Added anyway."

    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
